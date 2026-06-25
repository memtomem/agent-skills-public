"""Generate messy-spreadsheet fixtures (never committed; built on demand).

Builds messy.xlsx exercising every hard case:
  - Sheet "Finance": a title row + note above the real table (header not row 1,
    separated by a BLANK gutter), plus a SECOND stacked table, KR+EN labels.
  - Sheet "Merged": merged header cells (colspan) — must render as HTML.
  - Sheet "Formulas": a column of SUM formulas with NO cached values.
  - Sheet "NoGutter": a title sitting DIRECTLY above the header with no blank
    row between — the header-split edge case (finding #1).
  - Sheet "Chart": a small table + a BarChart referencing it, so chart category
    and series values can be resolved exactly (finding #3).
  - Sheet "Empty": blank.
"""
import os
import sys

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font


def _finance(ws):
    ws["A1"] = "2025 재무 요약 / Financial Summary"   # title strip
    ws["A2"] = "단위: 억원 (KRW 100M)"                  # note strip
    # real table starts at row 4 (header not on row 1)
    rows = [["항목 Item", "2023", "2024", "2025"],
            ["매출 Revenue", 980, 1050, 1239],
            ["영업이익 Op.Income", 88, 102, 154],
            ["순이익 Net Income", 61, 70, 112]]
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            ws.cell(row=4 + i, column=1 + j, value=v)
    # blank gutter rows 9-10, then a SECOND stacked table
    second = [["부문 Segment", "비중%"],
              ["국내 Domestic", 62],
              ["해외 Overseas", 38]]
    for i, row in enumerate(second):
        for j, v in enumerate(row):
            ws.cell(row=11 + i, column=1 + j, value=v)


def _merged(ws):
    ws["A1"] = "Region"
    ws["B1"] = "H1"            # will span B1:C1
    ws.merge_cells("B1:C1")
    ws["A2"] = "Quarter"
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    data = [["Asia", 120, 135], ["EU", 90, 110]]
    for i, row in enumerate(data):
        for j, v in enumerate(row):
            ws.cell(row=3 + i, column=1 + j, value=v)


def _formulas(ws):
    ws["A1"] = "Item"; ws["B1"] = "Qty"; ws["C1"] = "Price"; ws["D1"] = "Total"
    rows = [["Widget", 10, 5], ["Gadget", 4, 12]]
    for i, (name, qty, price) in enumerate(rows):
        r = 2 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=3, value=price)
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}")  # formula, no Excel cache


def _nogutter(ws):
    # Title strip DIRECTLY above the header (no blank-row gutter between them).
    ws["A1"] = "Q3 Sales by Region — preliminary"
    rows = [["Region", "Units", "Revenue"],
            ["Asia", 120, 4800],
            ["EU", 90, 3600],
            ["US", 75, 3000]]
    for i, row in enumerate(rows):
        for j, v in enumerate(row):
            ws.cell(row=2 + i, column=1 + j, value=v)  # header on row 2


def _chart(ws):
    ws["A1"] = "Quarter"; ws["B1"] = "Sales"
    data = [["Q1", 260], ["Q2", 295], ["Q3", 330], ["Q4", 354]]
    for i, (q, v) in enumerate(data):
        ws.cell(row=2 + i, column=1, value=q)
        ws.cell(row=2 + i, column=2, value=v)
    chart = BarChart()
    chart.title = "Quarterly Sales"
    values = Reference(ws, min_col=2, min_row=1, max_row=5)   # B1 title + B2:B5
    cats = Reference(ws, min_col=1, min_row=2, max_row=5)     # A2:A5
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")


def _allformula(ws):
    # A region whose cells are ALL uncached formulas — no literals, no cache.
    # With data_only detection this region would vanish entirely.
    ws["A1"] = "=1+2"; ws["B1"] = "=10*2"
    ws["A2"] = "=3+4"; ws["B2"] = "=10*3"


def build(path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Finance"; _finance(ws)
    _merged(wb.create_sheet("Merged"))
    _formulas(wb.create_sheet("Formulas"))
    _nogutter(wb.create_sheet("NoGutter"))
    _chart(wb.create_sheet("Chart"))
    _allformula(wb.create_sheet("AllFormula"))
    wb.create_sheet("Empty")
    wb.save(path)
    return path


def make_all(out_dir: str) -> None:
    os.makedirs(out_dir, exist_ok=True)
    build(os.path.join(out_dir, "messy.xlsx"))


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.dirname(__file__)
    make_all(out)
    print("fixtures written to", out)
