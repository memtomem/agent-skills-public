"""Core library for parsing messy, unstructured spreadsheets (.xlsx/.xlsm).

A spreadsheet file is *structured* (it's an OOXML zip), but it is rarely a clean
single table. Real sheets stack several tables with blank-row gutters, put the
header three rows down under a title and a note, merge cells for visual grouping,
hold formulas whose values you actually want, and embed charts. A naive
"read every cell into a dataframe" flattens all of that into garbage.

So this library follows a *detect-then-extract* strategy, mirroring the
pdf-parser philosophy:

1. For each sheet, find the rectangular **table regions** — blocks of populated
   cells separated by fully-empty rows/columns (``detect_table_regions``).
2. Within a region, find the **header row** heuristically (it isn't always row
   one) and read the body using **cached values** for formulas.
3. Preserve **merged cells** faithfully by rendering regions with merges as
   HTML (rowspan/colspan), which a flat pipe table cannot express.
4. Pull out **charts** (their data lives in the XML, so it's exact, not OCR).
5. Score each sheet's extraction **confidence** with reason flags so the risky
   sheets (ambiguous header, many merges, several tables) surface for review.

Output is the same shape as pdf-parser: a Markdown rendering plus a structured
JSON element tree, so both feed an LLM/RAG pipeline or downstream code.

Requires openpyxl. markitdown is used only as an optional cross-check.
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field, asdict
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------
# Element model (shared shape with the pdf-parser element tree)
# --------------------------------------------------------------------------

@dataclass
class Element:
    type: str                          # heading | paragraph | table | chart | note
    sheet: str
    text: str = ""
    rows: Optional[list[list[str]]] = None      # table: cell values
    spans: Optional[list[dict]] = None          # table: merged-cell spans
    ref: Optional[str] = None                    # A1-style region/anchor reference
    note: Optional[str] = None
    chart: Optional[dict] = None                 # chart metadata

    def to_dict(self) -> dict:
        d = asdict(self)
        return {k: v for k, v in d.items() if v not in (None, [], "")}


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------

class XlsxError(Exception):
    """User-facing error for an unreadable/missing workbook."""


def load_dual(path: str):
    """Return (wb_values, wb_formulas).

    openpyxl can't give both the formula and its cached result from one load,
    so we load twice: data_only=True yields the last-computed values (what you
    usually want), data_only=False yields the formula strings (for provenance).
    """
    if not os.path.exists(path):
        raise XlsxError(f"file not found: {path!r}")
    try:
        wb_v = openpyxl.load_workbook(path, data_only=True, read_only=False)
        wb_f = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as e:
        raise XlsxError(f"cannot open {path!r}: {e}") from e
    return wb_v, wb_f


# --------------------------------------------------------------------------
# Cell helpers
# --------------------------------------------------------------------------

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return re.sub(r"\s+", " ", str(v)).strip()


def _is_empty_row(ws, r: int, c0: int, c1: int) -> bool:
    return all(ws.cell(row=r, column=c).value in (None, "") for c in range(c0, c1 + 1))


def _is_empty_col(ws, c: int, r0: int, r1: int) -> bool:
    return all(ws.cell(row=r, column=c).value in (None, "") for r in range(r0, r1 + 1))


# --------------------------------------------------------------------------
# Table-region detection
# --------------------------------------------------------------------------

@dataclass
class Region:
    min_row: int
    min_col: int
    max_row: int
    max_col: int

    @property
    def ref(self) -> str:
        return f"{get_column_letter(self.min_col)}{self.min_row}:{get_column_letter(self.max_col)}{self.max_row}"


def detect_table_regions(ws) -> list[Region]:
    """Split a sheet into rectangular populated blocks.

    Algorithm: find the used bounding box, then recursively cut it on any fully
    empty row or column. This separates stacked tables (empty-row gutter) and
    side-by-side tables (empty-column gutter) without assuming the sheet holds
    just one table starting at A1 — the common reason a flat read goes wrong.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    def split(r0: int, r1: int, c0: int, c1: int) -> list[Region]:
        # Trim leading/trailing empty rows & cols.
        while r0 <= r1 and _is_empty_row(ws, r0, c0, c1):
            r0 += 1
        while r1 >= r0 and _is_empty_row(ws, r1, c0, c1):
            r1 -= 1
        while c0 <= c1 and _is_empty_col(ws, c0, r0, r1):
            c0 += 1
        while c1 >= c0 and _is_empty_col(ws, c1, r0, r1):
            c1 -= 1
        if r0 > r1 or c0 > c1:
            return []
        # Find an interior empty row to split on (horizontal cut).
        for r in range(r0 + 1, r1):
            if _is_empty_row(ws, r, c0, c1):
                return split(r0, r - 1, c0, c1) + split(r + 1, r1, c0, c1)
        # Find an interior empty column to split on (vertical cut).
        for c in range(c0 + 1, c1):
            if _is_empty_col(ws, c, r0, r1):
                return split(r0, r1, c0, c - 1) + split(r0, r1, c + 1, c1)
        return [Region(r0, c0, r1, c1)]

    return split(1, max_row, 1, max_col)


# --------------------------------------------------------------------------
# Header detection
# --------------------------------------------------------------------------

def _looks_numeric(v: Any) -> bool:
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        return bool(re.fullmatch(r"[-+]?[\d,]*\.?\d+%?", v.strip()))
    return False


def detect_header_row(ws, region: Region) -> tuple[int, bool]:
    """Return (header_row_index, confident).

    Heuristic: the header is the first row in the region that is mostly text
    while the row beneath it is markedly more numeric — the classic "labels on
    top of numbers" shape. If no row fits, fall back to the region's first row
    and report low confidence so the caller can flag it.
    """
    r0, r1, c0, c1 = region.min_row, region.max_row, region.min_col, region.max_col
    for r in range(r0, min(r0 + 5, r1)):
        cells = [ws.cell(row=r, column=c).value for c in range(c0, c1 + 1)]
        below = [ws.cell(row=r + 1, column=c).value for c in range(c0, c1 + 1)]
        nonempty = [v for v in cells if v not in (None, "")]
        if not nonempty:
            continue
        text_ratio = sum(0 if _looks_numeric(v) else 1 for v in nonempty) / len(nonempty)
        below_nonempty = [v for v in below if v not in (None, "")]
        below_num = (sum(1 for v in below_nonempty if _looks_numeric(v)) / len(below_nonempty)) if below_nonempty else 0
        # Path A: classic "text labels over numbers".
        if text_ratio >= 0.6 and below_num >= 0.4:
            return r, True
        # Path B: a matrix table whose column headers are themselves numeric
        # (years, codes). The tell is a fully-populated first row sitting atop a
        # body whose FIRST column is textual row-labels — then row 1 is headers
        # even though "2023/2024/2025" look numeric. Avoids over-flagging the
        # very common financial-table shape.
        if r == r0 and len(nonempty) == (c1 - c0 + 1) and r1 - r0 >= 2:
            label_col = [ws.cell(row=rr, column=c0).value for rr in range(r0 + 1, r1 + 1)]
            label_nonempty = [v for v in label_col if v not in (None, "")]
            if label_nonempty and sum(0 if _looks_numeric(v) else 1 for v in label_nonempty) / len(label_nonempty) >= 0.6:
                return r, True
    return r0, False


# --------------------------------------------------------------------------
# Merged cells & table extraction
# --------------------------------------------------------------------------

def _region_merges(ws, region: Region) -> list[tuple[int, int, int, int]]:
    """Merged ranges (1-based row/col, inclusive) intersecting the region."""
    out = []
    for mc in ws.merged_cells.ranges:
        if (mc.min_row >= region.min_row and mc.max_row <= region.max_row
                and mc.min_col >= region.min_col and mc.max_col <= region.max_col):
            out.append((mc.min_row, mc.min_col, mc.max_row, mc.max_col))
    return out


def extract_region(ws_v, ws_f, region: Region) -> tuple[list[list[str]], list[dict], bool, bool]:
    """Extract a region into (rows, spans, has_formula, missing_cache).

    rows: cached values as strings, full grid (merged followers blank). A formula
          cell whose cached result is absent (common when the file was written by
          a non-Excel tool) is surfaced as its **formula string** rather than a
          silent blank, and ``missing_cache`` is set so the caller can say so
          honestly instead of claiming "cached values shown".
    spans: list of {row, col, rowspan, colspan} (0-based within the region)
           for merged anchors — lets the renderer emit faithful HTML.
    """
    r0, r1, c0, c1 = region.min_row, region.max_row, region.min_col, region.max_col
    merges = _region_merges(ws_v, region)
    follower = set()
    span_map = {}
    for (mr0, mc0, mr1, mc1) in merges:
        span_map[(mr0, mc0)] = (mr1 - mr0 + 1, mc1 - mc0 + 1)
        for rr in range(mr0, mr1 + 1):
            for cc in range(mc0, mc1 + 1):
                if (rr, cc) != (mr0, mc0):
                    follower.add((rr, cc))

    rows: list[list[str]] = []
    spans: list[dict] = []
    has_formula = False
    missing_cache = False
    for r in range(r0, r1 + 1):
        row_vals = []
        for c in range(c0, c1 + 1):
            if (r, c) in follower:
                row_vals.append("")
                continue
            f = ws_f.cell(row=r, column=c).value
            v = ws_v.cell(row=r, column=c).value
            if isinstance(f, str) and f.startswith("="):
                has_formula = True
                if v in (None, ""):
                    # No cached result — show the formula itself (provenance)
                    # instead of a misleading blank.
                    missing_cache = True
                    cell_text = _cell_str(f)
                else:
                    cell_text = _cell_str(v)
            else:
                cell_text = _cell_str(v)
            row_vals.append(cell_text)
            if (r, c) in span_map:
                rs, cs = span_map[(r, c)]
                if rs > 1 or cs > 1:
                    spans.append({"row": r - r0, "col": c - c0, "rowspan": rs, "colspan": cs})
        rows.append(row_vals)
    return rows, spans, has_formula, missing_cache


# --------------------------------------------------------------------------
# Charts
# --------------------------------------------------------------------------

def _ref_cache(ref_obj) -> Optional[list]:
    """Cached point values carried inside a num/str reference (Excel-authored
    files embed these; openpyxl-authored ones usually don't)."""
    if ref_obj is None:
        return None
    cache = getattr(ref_obj, "numCache", None) or getattr(ref_obj, "strCache", None)
    pts = getattr(cache, "pt", None) if cache is not None else None
    if not pts:
        return None
    ordered = sorted(pts, key=lambda p: getattr(p, "idx", 0))
    return [p.v for p in ordered]


def _resolve_ref(wb_v, ref: Optional[str]) -> list:
    """Resolve a ``Sheet!$A$1:$A$9`` reference to a flat list of cell values."""
    if not ref or "!" not in ref:
        return []
    sheet_part, rng = ref.rsplit("!", 1)
    sheet_name = sheet_part.strip()
    if sheet_name.startswith("'") and sheet_name.endswith("'"):
        sheet_name = sheet_name[1:-1].replace("''", "'")
    if wb_v is None or sheet_name not in wb_v.sheetnames:
        return []
    ws = wb_v[sheet_name]
    try:
        block = ws[rng.replace("$", "")]
    except Exception:
        return []
    vals = []
    if isinstance(block, tuple):
        for row in block:
            for cell in (row if isinstance(row, tuple) else (row,)):
                vals.append(cell.value)
    else:
        vals.append(block.value)
    return vals


def _part_ref(part) -> Optional[str]:
    """The formula string from a numRef/strRef container, if any."""
    if part is None:
        return None
    for attr in ("numRef", "strRef"):
        sub = getattr(part, attr, None)
        if sub is not None and getattr(sub, "f", None):
            return sub.f
    return None


def _part_values(wb_v, part) -> list:
    """Resolve a numRef/strRef container to values — embedded cache first, then
    by reading the referenced cells from the (data_only) workbook."""
    if part is None:
        return []
    for attr in ("numRef", "strRef"):
        sub = getattr(part, attr, None)
        if sub is not None:
            cached = _ref_cache(sub)
            if cached:
                return cached
    return _resolve_ref(wb_v, _part_ref(part))


def extract_charts(ws, wb_v=None) -> list[Element]:
    out = []
    for ch in getattr(ws, "_charts", []) or []:
        title = None
        try:
            if ch.title and ch.title.tx and ch.title.tx.rich:
                title = "".join(
                    run.t or "" for p in ch.title.tx.rich.p for run in (p.r or [])
                ).strip() or None
        except Exception:
            title = None

        series_refs = []
        series = []
        categories: list[str] = []
        try:
            for s in ch.series:
                val_part = getattr(s, "val", None)
                cat_part = getattr(s, "cat", None)
                val_ref = _part_ref(val_part)
                if val_ref:
                    series_refs.append(val_ref)
                values = _part_values(wb_v, val_part)
                # Series name: from s.tx (a numRef/strRef, or a literal .v).
                sname = None
                tx = getattr(s, "tx", None)
                if tx is not None:
                    tref = _part_ref(tx)
                    if tref:
                        rv = _resolve_ref(wb_v, tref)
                        sname = _cell_str(rv[0]) if rv else None
                    elif getattr(tx, "v", None):
                        sname = _cell_str(tx.v)
                series.append({"name": sname, "ref": val_ref,
                               "values": [_cell_str(v) for v in values if v is not None]})
                # Categories are shared across series — capture the first non-empty.
                if not categories:
                    cvals = _part_values(wb_v, cat_part)
                    if cvals:
                        categories = [_cell_str(v) for v in cvals if v is not None]
        except Exception:
            pass

        chart = {"kind": ch.__class__.__name__, "title": title,
                 "series_refs": [r for r in series_refs if r]}
        if categories:
            chart["categories"] = categories
        if any(sd.get("values") for sd in series):
            chart["series"] = series
        out.append(Element(type="chart", sheet=ws.title,
                           text=title or f"{ch.__class__.__name__}", chart=chart))
    return out


# --------------------------------------------------------------------------
# Sheet + document assembly
# --------------------------------------------------------------------------

CONFIDENCE_VERIFY_BELOW = 0.75


def parse_sheet(ws_v, ws_f, wb_v=None) -> dict:
    name = ws_v.title
    # Detect populated regions against the FORMULA workbook: with data_only=True
    # an all-formula block with no cached results reads as empty and would be
    # dropped before extract_region can surface the formula strings. ws_f reports
    # both literals and formula text as populated, so such regions survive; values
    # are still pulled from ws_v inside extract_region.
    regions = detect_table_regions(ws_f)
    elements: list[Element] = []
    flags: list[str] = []
    confidence = 1.0

    for reg in regions:
        rows, spans, has_formula, missing_cache = extract_region(ws_v, ws_f, reg)
        if not rows:
            continue
        hdr_idx, header_ok = detect_header_row(ws_v, reg)
        # A 1-row or 1-col region is more likely a title/label strip than a table.
        if len(rows) == 1 or all(len(r) <= 1 for r in rows):
            txt = " ".join(c for r in rows for c in r if c)
            if txt:
                elements.append(Element(type="note", sheet=name, text=txt, ref=reg.ref))
            continue
        # If a confident header sits BELOW the region's first row, the rows above
        # it MAY be a title/note preamble swept in without a blank-row gutter.
        # Only split when every such row is a single-cell "strip" (the title/note
        # shape) — a multi-cell pre-header row is more likely the top of a
        # multi-row header (e.g. a merged super-header) and must be preserved.
        # (Finding #1: don't discard the detected header index — but don't be so
        # eager that a two-row header gets torn apart, either.)
        ref = reg.ref
        pre = hdr_idx - reg.min_row
        pre_is_strip = pre > 0 and all(sum(1 for c in r if c) <= 1 for r in rows[:pre])
        if header_ok and pre_is_strip:
            pre_txt = " ".join(c for r in rows[:pre] for c in r if c)
            if pre_txt:
                pre_ref = (f"{get_column_letter(reg.min_col)}{reg.min_row}:"
                           f"{get_column_letter(reg.max_col)}{hdr_idx - 1}")
                elements.append(Element(type="note", sheet=name, text=pre_txt, ref=pre_ref))
            rows = rows[pre:]
            spans = [{**s, "row": s["row"] - pre} for s in spans if s["row"] >= pre]
            ref = (f"{get_column_letter(reg.min_col)}{hdr_idx}:"
                   f"{get_column_letter(reg.max_col)}{reg.max_row}")
        if not rows:
            continue
        el = Element(type="table", sheet=name, rows=rows, ref=ref)
        if spans:
            el.spans = spans
        notes = []
        if has_formula:
            if missing_cache:
                notes.append("contains-formula(no cached values — "
                             "formulas shown literally; recalc in Excel/LibreOffice)")
            else:
                notes.append("contains-formula(cached values shown)")
        if not header_ok:
            notes.append("header-ambiguous")
        if notes:
            el.note = "; ".join(notes)
        elements.append(el)

    charts = extract_charts(ws_v, wb_v)
    elements.extend(charts)

    # Confidence flags.
    table_els = [e for e in elements if e.type == "table"]
    if len(table_els) > 1:
        confidence = min(confidence, 0.80)
        flags.append("multiple-table-regions")
    if any(e.spans for e in table_els):
        confidence = min(confidence, 0.75)
        flags.append("merged-cells")
    if any(e.note and "header-ambiguous" in e.note for e in table_els):
        confidence = min(confidence, 0.70)
        flags.append("header-ambiguous")
    if any(e.note and "contains-formula" in (e.note or "") for e in table_els):
        confidence = min(confidence, 0.85)
        flags.append("formula-cells")
    if any(e.note and "no cached values" in (e.note or "") for e in table_els):
        confidence = min(confidence, 0.65)
        flags.append("formula-no-cache")
    if not elements:
        confidence = min(confidence, 0.50)
        flags.append("empty-sheet")

    return {
        "name": name,
        "table_count": len(table_els),
        "confidence": round(confidence, 2),
        "flags": flags,
        "elements": [e.to_dict() for e in elements],
    }


def build_document(path: str) -> dict:
    wb_v, wb_f = load_dual(path)
    sheets = []
    for name in wb_v.sheetnames:
        sheets.append(parse_sheet(wb_v[name], wb_f[name], wb_v))
    min_conf = round(min((s["confidence"] for s in sheets), default=1.0), 2)
    return {
        "source": os.path.basename(path),
        "sheet_count": len(sheets),
        "min_confidence": min_conf,
        "verify_sheets": [s["name"] for s in sheets if s["confidence"] < CONFIDENCE_VERIFY_BELOW],
        "sheets": sheets,
    }


# --------------------------------------------------------------------------
# Rendering
# --------------------------------------------------------------------------

def _esc(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _gfm_cell(s: str) -> str:
    # HTML-escape (rendered-to-HTML safety) AND neutralize pipe/newline so the
    # cell can't break the surrounding GFM pipe-table layout.
    return _esc(s).replace("|", "\\|").replace("\n", " ")


def _rows_to_gfm(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    def fmt(r):
        return "| " + " | ".join(_gfm_cell(c) for c in r) + " |"
    out = [fmt(norm[0]), "| " + " | ".join(["---"] * width) + " |"]
    out += [fmt(r) for r in norm[1:]]
    return "\n".join(out)


def _rows_to_html(rows: list[list[str]], spans: Optional[list[dict]]) -> str:
    """Render with merged-cell rowspan/colspan and skip covered followers."""
    spans = spans or []
    anchor = {(s["row"], s["col"]): (s.get("rowspan", 1), s.get("colspan", 1)) for s in spans}
    covered = set()
    for s in spans:
        for dr in range(s.get("rowspan", 1)):
            for dc in range(s.get("colspan", 1)):
                if dr or dc:
                    covered.add((s["row"] + dr, s["col"] + dc))
    out = ["<table>"]
    for ri, row in enumerate(rows):
        tag = "th" if ri == 0 else "td"
        cells = []
        for ci, val in enumerate(row):
            if (ri, ci) in covered:
                continue
            attr = ""
            if (ri, ci) in anchor:
                rs, cs = anchor[(ri, ci)]
                if rs > 1:
                    attr += f' rowspan="{rs}"'
                if cs > 1:
                    attr += f' colspan="{cs}"'
            cells.append(f"<{tag}{attr}>{_esc(val).replace(chr(10), ' ')}</{tag}>")
        out.append("  <tr>" + "".join(cells) + "</tr>")
    out.append("</table>")
    return "\n".join(out)


def element_to_markdown(el: dict) -> str:
    # User-authored text is HTML-escaped on the way into Markdown (it may be
    # rendered back to HTML downstream); JSON keeps the raw values.
    t = el.get("type")
    text = _esc((el.get("text") or "").strip())
    if t in ("heading",):
        return f"## {text}"
    if t in ("note", "paragraph"):
        return text
    if t == "chart":
        ch = el.get("chart") or {}
        head = f"**[chart]** {text} _(kind: {_esc(str(ch.get('kind', '?')))})_"
        cats = [_gfm_cell(c) for c in (ch.get("categories") or [])]
        series = ch.get("series") or []
        if series:
            # Resolved data — render as a compact table (categories as columns).
            lines = [head]
            if cats:
                lines.append("| series | " + " | ".join(cats) + " |")
                lines.append("| --- | " + " | ".join(["---"] * len(cats)) + " |")
                for s in series:
                    vals = [_gfm_cell(v) for v in (s.get("values") or [])] + [""] * len(cats)
                    name = _gfm_cell(s.get("name") or "series")
                    lines.append("| " + name + " | " + " | ".join(vals[:len(cats)]) + " |")
            else:
                for s in series:
                    vals = ", ".join(_esc(v) for v in (s.get("values") or [])) or "n/a"
                    lines.append(f"- {_esc(s.get('name') or 'series')}: {vals}")
            return "\n".join(lines)
        refs = ", ".join(_esc(r) for r in (ch.get("series_refs") or [])) or "n/a"
        return f"{head} data refs: {refs}"
    if t == "table":
        rows = el.get("rows") or []
        spans = el.get("spans")
        body = _rows_to_html(rows, spans) if spans else _rows_to_gfm(rows)
        note = el.get("note")
        ref = el.get("ref")
        head = f"<!-- table {ref}" + (f" · {note}" if note else "") + " -->"
        return head + "\n" + body
    return el.get("text", "").strip()


def to_markdown(document: dict) -> str:
    out = [f"<!-- parsed from {_esc(document.get('source', ''))} -->\n"]
    for sh in document.get("sheets", []):
        name = _esc(sh["name"])
        out.append(f"# Sheet: {name}")
        conf = sh.get("confidence")
        if conf is not None and conf < CONFIDENCE_VERIFY_BELOW:
            out.append(f"> 🔎 Sheet '{name}' low parse confidence ({conf}); "
                       f"verify. flags: {', '.join(sh.get('flags') or []) or 'none'}")
        for el in sh.get("elements", []):
            md = element_to_markdown(el)
            if md:
                out.append(md)
        out.append("")
    return "\n\n".join(out).strip() + "\n"


def save_outputs(document: dict, md_path: str, json_path: str) -> None:
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(document, f, ensure_ascii=False, indent=2)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(to_markdown(document))


# --------------------------------------------------------------------------
# Optional markitdown cross-check
# --------------------------------------------------------------------------

def markitdown_crosscheck(path: str) -> Optional[str]:
    """Return markitdown's markdown for the workbook, or None if unavailable.

    Used only as a second opinion — handy to diff against our structured output
    when a sheet looks wrong. We don't depend on it.
    """
    try:
        from markitdown import MarkItDown
    except Exception:
        return None
    try:
        return MarkItDown().convert(path).text_content
    except Exception:
        return None
